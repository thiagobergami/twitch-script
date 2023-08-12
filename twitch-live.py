import requests
import openpyxl
from datetime import datetime
import os

# Twitch API endpoint to get all categories
TWITCH_CATEGORIES_API = "https://api.twitch.tv/helix/games/top"
# Twitch API endpoint to get live streams by game ID
TWITCH_LIVE_STREAMS_API = "https://api.twitch.tv/helix/streams"

# Twitch Client ID - Replace 'YOUR_TWITCH_CLIENT_ID' with your actual Twitch client ID
TWITCH_CLIENT_ID = "dwn4nnujrqkg7dlrte188o1so7am8y"
TWITCH_CLIENT_SECRET = "y3f20sh2nyvu0fbo73lz9rwyvkumer"

# Number of top categories to retrieve
NUM_CATEGORIES = 100  # You can change this value to get more or fewer categories

def get_oauth_token():
    oauth_url = "https://id.twitch.tv/oauth2/token"
    data = {
        "client_id": TWITCH_CLIENT_ID,
        "client_secret": TWITCH_CLIENT_SECRET,
        "grant_type": "client_credentials",
    }
    response = requests.post(oauth_url, data=data)
    response_json = response.json()

    if "access_token" in response_json:
        return response_json["access_token"]
    else:
        return None

def get_top_twitch_categories(oauth_token):
    headers = {"Client-ID": TWITCH_CLIENT_ID, "Authorization": f"Bearer {oauth_token}"}
    params = {"first": NUM_CATEGORIES}
    categories_data = []

    count_categories = 0
    while count_categories < 10:
        response = requests.get(TWITCH_CATEGORIES_API, headers=headers, params=params)
        response_json = response.json()

        if "data" in response_json:
            categories_data.extend(response_json["data"])

        # Check if there's another page (pagination)
        if "pagination" in response_json and "cursor" in response_json["pagination"]:
            params["after"] = response_json["pagination"]["cursor"]
        else:
            break
        count_categories += 1

    return categories_data

def get_live_viewers_and_channels_by_category(category_id, oauth_token):
    headers = {"Client-ID": TWITCH_CLIENT_ID, "Authorization": f"Bearer {oauth_token}"}
    params = {"game_id": category_id, "first": NUM_CATEGORIES}

    response = requests.get(TWITCH_LIVE_STREAMS_API, headers=headers, params=params)
    response_json = response.json()

    total_live_viewers = 0
    total_live_channels = 0

    while True:
        response = requests.get(TWITCH_LIVE_STREAMS_API, headers=headers, params=params)
        response_json = response.json()

        if "data" in response_json:
            for stream in response_json["data"]:
                total_live_viewers += stream["viewer_count"]
                total_live_channels += 1

        # Check if there's another page (pagination)
        if "pagination" in response_json and "cursor" in response_json["pagination"]:
            params["after"] = response_json["pagination"]["cursor"]
        else:
            break

    return total_live_viewers, total_live_channels

def create_excel_file(categories_data, oauth_token):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Twitch Categories Data"

    sheet["A1"] = "Category"
    sheet["B1"] = "Total Live Viewers"
    sheet["C1"] = "Total Live Channels"

    for row_index, category_data in enumerate(categories_data, start=2):
        category_name = category_data["name"]
        total_live_viewers, total_live_channels = get_live_viewers_and_channels_by_category(category_data["id"], oauth_token)

        print("Category name: " + category_name + "/ Total views: "+ str(total_live_viewers) + "/ total_streams: "+ str(total_live_channels))

        sheet.cell(row=row_index, column=1, value=category_name)
        sheet.cell(row=row_index, column=2, value=total_live_viewers)
        sheet.cell(row=row_index, column=3, value=total_live_channels)

    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"{today_date}_live_data.xlsx"
    data_folder = "data"
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)

    full_file_path = os.path.join(data_folder, excel_filename)
    workbook.save(full_file_path)
    print(f"Data saved to '{full_file_path}'.")

if __name__ == "__main__":
    oauth_token = get_oauth_token()
    if not oauth_token:
        print("Failed to get the OAuth token. Please check your client ID and client secret.")
    else:
        categories_data = get_top_twitch_categories(oauth_token)
        if categories_data:
            create_excel_file(categories_data, oauth_token)
        else:
            print("No data found.")